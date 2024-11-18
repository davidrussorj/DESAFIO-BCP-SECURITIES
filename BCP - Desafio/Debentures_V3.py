import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Carregar o arquivo Excel com os dados
df = pd.read_excel("DataSet v2.xlsx")

# Selecionar e ajustar colunas
df_selecionado = df[["Código", "Nome", "Índice/ Correção", "Taxa de Compra", "Taxa de Venda", "Taxa Indicativa", "PU", "Indexador", "Data"]]
df_selecionado['Taxa Indicativa'] = df_selecionado['Taxa Indicativa'].replace('--', np.nan)
df_selecionado['Data'] = pd.to_datetime(df_selecionado['Data']).dt.date  # Certificar que a coluna Data é do tipo date
df_selecionado['Indexador'] = df_selecionado['Indexador'].str.replace('0', '', regex=False)

# Criar arquivo base para exportação
output_path = 'DataSet_Final_Research.xlsx'
df_selecionado.to_excel(output_path, index=False)

# Abrir arquivo criado
workbook = load_workbook(output_path)

# Remover a planilha inicial padrão (opcional, se for reutilizável)
default_sheet = workbook.active
workbook.remove(default_sheet)

# Criar uma aba para cada data distinta
unique_dates = df_selecionado['Data'].unique()

for date in unique_dates:
    sheet_name = date.strftime('%Y-%m-%d')  # Formatar nome da aba com a data
    filtered_data = df_selecionado[df_selecionado['Data'] == date]
    
    # Adicionar aba ao workbook
    worksheet = workbook.create_sheet(title=sheet_name)
    
    # Inserir cabeçalhos
    for col_num, column_title in enumerate(filtered_data.columns, 1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        # Aplicar estilo ao cabeçalho
        header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.font = header_font
    
    # Inserir dados
    for row_num, row_data in enumerate(filtered_data.itertuples(index=False), 2):
        for col_num, value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)
    
    # Ajustar largura das colunas automaticamente
    for column in worksheet.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:  # Verifica se há valor
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[col_letter].width = adjusted_width

# Salvar o arquivo com as alterações
workbook.save(output_path)

print(f"Arquivo atualizado salvo em {output_path}")

