import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Caminho da pasta onde estão os arquivos Excel
folder_path = 'C:/Users/david/OneDrive/Documents/BCP - Desafio/Daily Prices'

# Listas para armazenar os DataFrames de cada planilha
df1_list = []
df2_list = []
df3_list = []

# Função para ler e manipular as planilhas de cada arquivo Excel
def manipular_planilhas(caminho_arquivo):
    # Extrair a data do nome do arquivo
    nome_arquivo = os.path.basename(caminho_arquivo)
    data_extraida = re.search(r'(\d{8})', nome_arquivo)  # Captura 8 dígitos consecutivos

    if data_extraida:
        # Converter a data extraída para o formato datetime (YYYY-MM-DD)
        data_fixa = pd.to_datetime(data_extraida.group(1), format='%Y%m%d')
    else:
        # Se a data não for encontrada, você pode definir uma data padrão ou lançar um erro
        data_fixa = pd.to_datetime('2024-01-01')  # Exemplo de data padrão
    
    if data_extraida:
        # Converter a data extraída para o formato datetime
        data_fixa = pd.to_datetime(data_extraida.group(1))
    else:
        # Se a data não for encontrada, você pode definir uma data padrão ou lançar um erro
        data_fixa = pd.to_datetime('2024-01-01')  # Exemplo de data padrão
    # Carregar todas as planilhas de um arquivo Excel
    df1 = pd.read_excel(caminho_arquivo, sheet_name='DI_PERCENTUAL')
    df2 = pd.read_excel(caminho_arquivo, sheet_name='DI_SPREAD')
    df3 = pd.read_excel(caminho_arquivo, sheet_name='IPCA_SPREAD')

    # Remover as 7 primeiras linhas de cada DataFrame
    linhas_para_remover = [0,1,2,3,4, 5,7]
    df1 =  df1.drop(linhas_para_remover)
    df2 =  df2.drop(linhas_para_remover)
    df3 =  df3.drop(linhas_para_remover)
    df1['data'] = data_fixa
    df2['data'] = data_fixa
    df3['data'] = data_fixa
    
    # Adicionar os DataFrames à lista correspondente
    df1_list.append(df1)
    df2_list.append(df2)
    df3_list.append(df3)

    # Retornar os DataFrames modificados, se necessário
    return df1, df2, df3

# Função para ajustar a largura das colunas
def ajustar_largura_colunas(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Obtém a letra da coluna
        for cell in col:
            try:
                # Calcula o comprimento máximo do conteúdo na coluna
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Define a largura da coluna com base no comprimento máximo
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width
    
# Percorrer todos os arquivos Excel na pasta
for arquivo in os.listdir(folder_path):
    
        caminho_arquivo = os.path.join(folder_path, arquivo)
        print(caminho_arquivo)
        manipular_planilhas(caminho_arquivo)

# Concatenar todos os DataFrames das listas em um único DataFrame para cada planilha
df1_total = pd.concat(df1_list, ignore_index=True)
df2_total = pd.concat(df2_list, ignore_index=True)
df3_total = pd.concat(df3_list, ignore_index=True)
print(df1_total)
import pandas as pd

# Função para salvar os DataFrames concatenados em um único arquivo Excel e ajustar as larguras
def salvar_arquivo_com_ajuste(df1_total, df2_total, df3_total, caminho_arquivo):
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        # Escrever cada DataFrame em uma sheet separada
        df1_total.to_excel(writer, sheet_name='%DI', index=False)
        df2_total.to_excel(writer, sheet_name='DI+', index=False)
        df3_total.to_excel(writer, sheet_name='IPCA+', index=False)

        # Obter o arquivo Excel gerado
        workbook = writer.book

        # Ajustar a largura das colunas para cada sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            ajustar_largura_colunas(sheet)

df1_total['Indexador'] = '%DI' 
df2_total['Indexador'] = 'DI+'
df3_total['Indexador'] = 'IPCA+'

salvar_arquivo_com_ajuste(df1_total, df2_total, df3_total, 'C:/Users/david/OneDrive/Documents/BCP - Desafio/DataSet V1.xlsx')

print(df1_total)