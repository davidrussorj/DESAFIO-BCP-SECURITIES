import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Caminho do arquivo Excel 
file_path = "DataSet V1.xlsx"

# Ler cada sheet individualmente com pandas
sheet1 = pd.read_excel(file_path, sheet_name='%DI',header=1)
sheet2 = pd.read_excel(file_path, sheet_name='DI+',header=1)
sheet3 = pd.read_excel(file_path, sheet_name='IPCA+',header=1)

# Concatenar as três sheets, ignorando o índice original
combined_df = pd.concat([sheet1, sheet2, sheet3], ignore_index=True)

# Excluir linhas onde o valor da primeira coluna é nulo
combined_df = combined_df.dropna(subset=[combined_df.columns[0]])

# Excluir linhas onde a primeira coluna começa com '( , Obs, Código'
combined_df = combined_df[~combined_df.iloc[:, 0].astype(str).str.startswith('(')]
combined_df = combined_df[~combined_df.iloc[:, 0].astype(str).str.startswith('Obs')]
combined_df = combined_df[~combined_df.iloc[:, 0].astype(str).str.startswith('Código')]

combined_df['%DI'] = combined_df['%DI'].fillna(0) 
combined_df['DI+'] = combined_df['DI+'].fillna(0)
combined_df['IPCA+'] = combined_df['IPCA+'].fillna(0)  
combined_df = combined_df.rename(columns={'Intervalo Indicativo': 'Min.', 'Unnamed: 9': 'Max'})

# Combine 'Column1' and 'Column2' into a new column 'Combined'
combined_df['Indexador'] = combined_df['%DI'].astype(str) + combined_df['DI+'].astype(str) + combined_df['IPCA+'].astype(str)
combined_df = combined_df.drop(columns=['%DI', 'DI+','IPCA+'])
combined_df['Indexador'] = combined_df['Indexador'].str.replace('0', '', regex=False)
# Ordenar o DataFrame pela coluna "Código" em ordem alfabética
combined_df = combined_df.sort_values(by='Código').reset_index(drop=True)

combined_df.columns.values[15] = 'Data'  

combined_df.to_excel("DataSet v2.xlsx", sheet_name='Sheet1', index=False)

print(combined_df.dtypes)

# Função para ajustar a largura das colunas
def ajustar_largura_colunas(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Obtém a letra da coluna
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width

# Reabrir o arquivo Excel com openpyxl para ajustar a largura das colunas
workbook = load_workbook("DataSet v2.xlsx")
sheet = workbook['Sheet1']
ajustar_largura_colunas(sheet)
workbook.save("DataSet v2.xlsx")

print("As sheets foram combinadas e ordenadas pelo 'Código', e as larguras das colunas ajustadas com sucesso.")
